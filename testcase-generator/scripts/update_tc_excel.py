#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ref_mapping.json을 기반으로 TC Excel의 Reference 컬럼을 업데이트하는 스크립트

- 원본 Excel 자동 백업
- 낮은 confidence → 노란 배경 + 코멘트
- --dry-run 모드 지원
"""

import json
import sys
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment

# 중앙 설정에서 가져오기
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try:
    from config import EXCEL_COLORS, REF_MAP_CONFIDENCE_THRESHOLD
except ImportError:
    EXCEL_COLORS = {"blank_field": "FFFF00"}
    REF_MAP_CONFIDENCE_THRESHOLD = 0.7


# 색상 정의
YELLOW_FILL = PatternFill(
    start_color=EXCEL_COLORS["blank_field"],
    end_color=EXCEL_COLORS["blank_field"],
    fill_type="solid",
)


def find_header_row(sheet, max_rows: int = 30) -> Optional[int]:
    """헤더 행 찾기"""
    search_keywords = [
        "test case id", "tc id", "testcase id", "테스트케이스", "tc번호",
    ]
    for row_idx in range(1, max_rows + 1):
        for col_idx in range(1, 30):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if not cell_value:
                continue
            cell_lower = str(cell_value).lower().strip()
            for keyword in search_keywords:
                if keyword in cell_lower:
                    return row_idx
    return None


def find_reference_column(sheet, header_row: int) -> Optional[int]:
    """Reference 컬럼 찾기"""
    ref_keywords = ["reference", "참조", "ref"]
    for col_idx in range(1, 50):
        cell_value = sheet.cell(row=header_row, column=col_idx).value
        if not cell_value:
            continue
        cell_lower = str(cell_value).lower().strip()
        for keyword in ref_keywords:
            if keyword == cell_lower or keyword in cell_lower:
                return col_idx
    return None


def find_tc_id_column(sheet, header_row: int) -> Optional[int]:
    """TC ID 컬럼 찾기"""
    id_keywords = ["test case id", "tc id", "testcase id", "테스트케이스 id", "tc번호"]
    for col_idx in range(1, 50):
        cell_value = sheet.cell(row=header_row, column=col_idx).value
        if not cell_value:
            continue
        cell_lower = str(cell_value).lower().strip()
        for keyword in id_keywords:
            if keyword in cell_lower:
                return col_idx
    return None


def backup_excel(excel_path: Path) -> Path:
    """원본 Excel 백업"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{excel_path.stem}_backup_{timestamp}{excel_path.suffix}"
    backup_path = excel_path.parent / backup_name
    shutil.copy2(excel_path, backup_path)
    return backup_path


def update_tc_excel(
    excel_path: Path,
    mapping_path: Path,
    output_path: Optional[Path] = None,
    dry_run: bool = False,
    confidence_threshold: float = REF_MAP_CONFIDENCE_THRESHOLD,
) -> Dict[str, Any]:
    """TC Excel Reference 컬럼 업데이트

    Args:
        excel_path: 원본 TC Excel 경로
        mapping_path: ref_mapping.json 경로
        output_path: 출력 경로 (None이면 원본 경로에 덮어쓰기)
        dry_run: True면 실제 저장 없이 결과만 출력
        confidence_threshold: 이 값 미만이면 노란 배경 + 코멘트

    Returns:
        업데이트 통계
    """
    print("=" * 60)
    print("  TC Excel Reference 업데이트")
    print("=" * 60)

    if dry_run:
        print("  [DRY RUN] 실제 저장 없이 결과만 출력합니다.")

    # 매핑 데이터 로드
    print("\n매핑 데이터 로딩...")
    with open(mapping_path, "r", encoding="utf-8") as f:
        mapping_data = json.load(f)

    mappings = mapping_data.get("mappings", [])
    print(f"  총 {len(mappings)}개 매핑")

    # row_index → mapping 딕셔너리
    row_mapping = {}
    # tc_id → mapping 딕셔너리 (row_index 없는 경우 대비)
    id_mapping = {}
    for m in mappings:
        ref = m.get("reference", "")
        if not ref:
            continue
        if m.get("row_index"):
            row_mapping[m["row_index"]] = m
        if m.get("test_case_id"):
            id_mapping[m["test_case_id"]] = m

    # Excel 열기
    if not dry_run:
        backup_path = backup_excel(excel_path)
        print(f"백업: {backup_path}")

    wb = load_workbook(excel_path)
    sheet = wb.active

    # 헤더 찾기
    header_row = find_header_row(sheet)
    if header_row is None:
        print("Error: 헤더 행을 찾을 수 없습니다.")
        wb.close()
        sys.exit(1)

    ref_col = find_reference_column(sheet, header_row)
    tc_id_col = find_tc_id_column(sheet, header_row)

    if ref_col is None:
        # Reference 컬럼이 없으면 마지막 데이터 컬럼 뒤에 새로 생성
        # 마지막 사용 컬럼 찾기
        last_col = 1
        for col_idx in range(1, 50):
            if sheet.cell(row=header_row, column=col_idx).value:
                last_col = col_idx
        ref_col = last_col + 1
        sheet.cell(row=header_row, column=ref_col).value = "Reference"
        print(f"  Reference 컬럼 없음 → 새로 생성 (컬럼 {ref_col})")

    print(f"헤더 행: {header_row}, Reference 컬럼: {ref_col}")

    # 업데이트
    data_start = header_row + 1
    stats = {
        "updated": 0,
        "skipped_no_mapping": 0,
        "skipped_existing": 0,
        "low_confidence": 0,
    }

    for row_idx in range(data_start, sheet.max_row + 1):
        # TC ID 확인 (데이터 행인지 판별)
        if tc_id_col:
            tc_id_value = sheet.cell(row=row_idx, column=tc_id_col).value
            if not tc_id_value:
                continue
            tc_id_str = str(tc_id_value).strip()
        else:
            tc_id_str = ""

        # 매핑 찾기 (row_index 우선, 없으면 TC ID로)
        mapping = row_mapping.get(row_idx)
        if not mapping and tc_id_str:
            mapping = id_mapping.get(tc_id_str)

        if not mapping:
            stats["skipped_no_mapping"] += 1
            continue

        reference = mapping.get("reference", "")
        if not reference:
            stats["skipped_no_mapping"] += 1
            continue

        confidence = mapping.get("confidence", 0)
        reasoning = mapping.get("reasoning", "")

        # 기존 Reference 확인
        current_ref = sheet.cell(row=row_idx, column=ref_col).value
        if current_ref and str(current_ref).strip():
            # needs_mapping 여부는 tc_input.json에서 이미 필터링됨
            # 여기서는 매핑 결과가 있으면 업데이트
            pass

        # 셀 업데이트
        cell = sheet.cell(row=row_idx, column=ref_col)
        cell.value = reference
        stats["updated"] += 1

        # 낮은 confidence → 노란 배경 + 코멘트
        if confidence < confidence_threshold:
            cell.fill = YELLOW_FILL
            comment_text = (
                f"[자동매핑] 신뢰도: {confidence:.0%}\n"
                f"[근거] {reasoning}"
            )
            cell.comment = Comment(text=comment_text, author="Reference Mapper")
            stats["low_confidence"] += 1

    # 저장
    if not dry_run:
        save_path = output_path if output_path else excel_path
        wb.save(save_path)
        print(f"\n저장: {save_path}")
    else:
        print("\n[DRY RUN] 저장하지 않았습니다.")

    wb.close()

    print(f"\n업데이트 결과:")
    print(f"  업데이트: {stats['updated']}건")
    print(f"  매핑 없음 (건너뜀): {stats['skipped_no_mapping']}건")
    print(f"  낮은 신뢰도 (노란 배경): {stats['low_confidence']}건")
    print("=" * 60)

    return stats


def main():
    if len(sys.argv) < 3:
        print("Usage: python update_tc_excel.py <TC.xlsx> <ref_mapping.json> [options]")
        print()
        print("Options:")
        print("  --output <path>   출력 Excel 경로 (기본: 원본 덮어쓰기)")
        print("  --dry-run         실제 저장 없이 결과만 출력")
        sys.exit(1)

    excel_path = Path(sys.argv[1])
    mapping_path = Path(sys.argv[2])

    # 옵션 파싱
    output_path = None
    dry_run = False

    args = sys.argv[3:]
    i = 0
    while i < len(args):
        if args[i] == "--output" and i + 1 < len(args):
            output_path = Path(args[i + 1])
            i += 2
        elif args[i] == "--dry-run":
            dry_run = True
            i += 1
        else:
            i += 1

    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)

    if not mapping_path.exists():
        print(f"Error: File not found: {mapping_path}")
        sys.exit(1)

    update_tc_excel(excel_path, mapping_path, output_path, dry_run)


if __name__ == "__main__":
    main()
