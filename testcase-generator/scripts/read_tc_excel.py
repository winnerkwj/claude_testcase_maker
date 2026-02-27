#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
기존 TC Excel 파일을 읽어 tc_input.json으로 변환하는 스크립트

다양한 Excel 형식을 지원하며, Reference 매핑이 필요한 TC를 식별합니다.
- 동적 헤더 행 감지 (1~30행 스캔)
- 유연한 컬럼 매핑 (컬럼명 유사어 지원)
- 기존 Reference 보존 / 덮어쓰기 모드 지원
"""

import json
import sys
import re
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple

from openpyxl import load_workbook

# 중앙 설정에서 가져오기
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try:
    from config import OUTPUT_DIR
except ImportError:
    OUTPUT_DIR = Path(__file__).resolve().parent.parent.parent / "output"


# 컬럼명 유사어 매핑 (소문자 기준)
COLUMN_ALIASES = {
    "test_case_id": [
        "test case id", "tc id", "testcase id", "테스트케이스 id",
        "테스트케이스id", "tc_id", "testcaseid", "tc번호",
    ],
    "depth1": ["depth 1", "depth1", "대분류"],
    "depth2": ["depth 2", "depth2", "중분류"],
    "depth3": ["depth 3", "depth3", "소분류", "기능 영역"],
    "depth4": ["depth 4", "depth4", "조건", "상태"],
    "title": [
        "title", "제목", "테스트 제목", "tc 제목",
        "description",
    ],
    "tc_name": ["tc name"],
    "pre_condition": [
        "pre-condition", "precondition", "pre condition",
        "사전조건", "사전 조건", "전제조건",
    ],
    "test_step": [
        "test step", "test steps", "테스트 절차", "테스트 스텝",
        "steps", "절차", "수행 절차", "step action",
    ],
    "expected_result": [
        "expected result", "expected results", "기대결과", "기대 결과",
        "예상결과", "예상 결과", "step expected result", "기대되는",
    ],
    "requirement_id": [
        "요구사항 id", "요구사항id", "requirement id", "requirement",
        "요구사항", "req id", "srs id",
    ],
    "reference": [
        "reference", "참조", "ref", "페이지", "화면 참조",
    ],
    "importance": ["중요도", "importance", "priority", "우선순위"],
    "writer": ["writer", "작성자", "담당자"],
    "feature": ["feature", "기능"],
    "note": ["step note", "note", "비고"],
}


def find_header_row(sheet, max_rows: int = 30) -> Optional[int]:
    """헤더 행 찾기 (Test Case ID 또는 유사 키워드가 있는 행)

    1~max_rows 행을 스캔하며, TC ID 관련 키워드를 포함하는 셀을 찾습니다.
    """
    search_keywords = [
        "test case id", "tc id", "testcase id", "테스트케이스", "tc번호",
    ]

    for row_idx in range(1, max_rows + 1):
        for col_idx in range(1, 30):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if not cell_value:
                continue
            cell_lower = str(cell_value).lower().strip()
            for keyword in search_keywords:
                if keyword in cell_lower:
                    return row_idx

    return None


def find_column_mapping(sheet, header_row: int) -> Dict[str, int]:
    """헤더 행에서 컬럼 매핑 찾기 (유사어 지원)"""
    column_map = {}

    for col_idx in range(1, 50):
        cell_value = sheet.cell(row=header_row, column=col_idx).value
        if not cell_value:
            continue

        cell_lower = str(cell_value).lower().strip()

        for field_name, aliases in COLUMN_ALIASES.items():
            if field_name in column_map:
                continue
            for alias in aliases:
                if alias == cell_lower or alias in cell_lower:
                    column_map[field_name] = col_idx
                    break

    return column_map


def read_tc_excel(
    excel_path: Path,
    sheet_name: Optional[str] = None,
    overwrite: bool = False,
    output_path: Optional[Path] = None,
) -> Dict[str, Any]:
    """TC Excel 파일 읽기

    Args:
        excel_path: 입력 Excel 파일 경로
        sheet_name: 시트명 (None이면 활성 시트)
        overwrite: True면 기존 Reference도 재매핑 대상으로 포함
        output_path: 출력 JSON 경로

    Returns:
        tc_input 딕셔너리
    """
    print("=" * 60)
    print("  TC Excel 읽기")
    print("=" * 60)

    # read_only=False로 로드하여 전체 메모리 적재 (개별 셀 접근보다 훨씬 빠름)
    wb = load_workbook(excel_path, read_only=False, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
    print(f"시트: {sheet.title}")

    # 헤더 행 감지
    header_row = find_header_row(sheet)
    if header_row is None:
        print("Error: 헤더 행을 찾을 수 없습니다.")
        print("  'Test Case ID', 'TC ID', '테스트케이스' 등의 키워드가 포함된 행이 필요합니다.")
        print("  1~30행 범위에서 검색했으나 발견되지 않았습니다.")
        wb.close()
        sys.exit(1)

    print(f"헤더 행: {header_row}")

    # 컬럼 매핑
    column_map = find_column_mapping(sheet, header_row)
    print(f"컬럼 매핑: {column_map}")

    if "test_case_id" not in column_map:
        print("Error: 'Test Case ID' 컬럼을 찾을 수 없습니다.")
        wb.close()
        sys.exit(1)

    # 데이터 읽기
    data_start_row = header_row + 1
    testcases = []
    total_count = 0
    needs_mapping_count = 0

    tc_id_col = column_map.get("test_case_id")
    ref_col = column_map.get("reference")

    # 필요한 컬럼 인덱스 목록 (0-based로 변환)
    max_col_needed = max(column_map.values())

    for row in sheet.iter_rows(min_row=data_start_row, max_row=sheet.max_row,
                                min_col=1, max_col=max_col_needed, values_only=False):
        # row는 셀 튜플, 1-based column → 0-based index
        tc_id_cell = row[tc_id_col - 1] if tc_id_col - 1 < len(row) else None
        tc_id_value = tc_id_cell.value if tc_id_cell else None
        if not tc_id_value:
            continue
        tc_id_str = str(tc_id_value).strip()
        if not tc_id_str:
            continue

        total_count += 1
        row_idx = tc_id_cell.row

        # 각 필드 읽기
        tc = {
            "row_index": row_idx,
            "test_case_id": tc_id_str,
        }

        for field_name, col_idx in column_map.items():
            if field_name == "test_case_id":
                continue
            cell = row[col_idx - 1] if col_idx - 1 < len(row) else None
            cell_value = cell.value if cell else None
            tc[field_name] = str(cell_value).strip() if cell_value else ""

        # Reference 매핑 필요 여부 판단
        current_ref = tc.get("reference", "")
        tc["current_reference"] = current_ref

        if overwrite:
            tc["needs_mapping"] = True
            needs_mapping_count += 1
        else:
            if not current_ref:
                tc["needs_mapping"] = True
                needs_mapping_count += 1
            else:
                tc["needs_mapping"] = False

        testcases.append(tc)

    wb.close()

    # 결과 구성
    result = {
        "source_file": str(excel_path.resolve()),
        "header_row": header_row,
        "column_mapping": column_map,
        "total_tcs": total_count,
        "tcs_needing_mapping": needs_mapping_count,
        "overwrite_mode": overwrite,
        "testcases": testcases,
    }

    # 저장
    if output_path is None:
        output_path = OUTPUT_DIR / "tc_input.json"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\n총 TC 수: {total_count}")
    print(f"매핑 필요: {needs_mapping_count}")
    print(f"출력 파일: {output_path}")
    print("=" * 60)

    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python read_tc_excel.py <TC.xlsx> [options]")
        print()
        print("Options:")
        print("  --output <path>   출력 JSON 경로 (기본: output/tc_input.json)")
        print("  --overwrite       기존 Reference도 재매핑 대상으로 포함")
        print("  --sheet <name>    시트명 지정")
        sys.exit(1)

    excel_path = Path(sys.argv[1])

    # 옵션 파싱
    output_path = None
    overwrite = False
    sheet_name = None

    args = sys.argv[2:]
    i = 0
    while i < len(args):
        if args[i] == "--output" and i + 1 < len(args):
            output_path = Path(args[i + 1])
            i += 2
        elif args[i] == "--overwrite":
            overwrite = True
            i += 1
        elif args[i] == "--sheet" and i + 1 < len(args):
            sheet_name = args[i + 1]
            i += 2
        else:
            i += 1

    if not excel_path.exists():
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)

    read_tc_excel(excel_path, sheet_name, overwrite, output_path)


if __name__ == "__main__":
    main()
