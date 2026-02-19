#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
테스트케이스를 Excel 템플릿에 작성하는 스크립트
"""

import json
import sys
import shutil
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

# 중앙 설정에서 가져오기
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
try:
    from config import (
        EXCEL_COLORS,
        RESULT_OPTIONS,
        SEVERITY_OPTIONS,
        BASE_COLUMNS,
        ROUND_SUBCOLUMNS,
        NUM_TEST_ROUNDS,
    )
except ImportError:
    # config.py를 찾을 수 없는 경우 기본값 사용
    EXCEL_COLORS = {
        "header": "4472C4",
        "subheader": "5B9BD5",
        "summary": "D9E2F3",
        "round1": "70AD47",
        "round2": "FFC000",
        "round3": "ED7D31",
        "blank_field": "FFFF00",
    }
    RESULT_OPTIONS = ["Pass", "Fail", "N/T", "Block"]
    SEVERITY_OPTIONS = ["Critical", "Major", "Minor", "Trivial"]
    BASE_COLUMNS = [
        ("No", 5), ("Test Case ID", 15), ("Depth 1", 12), ("Depth 2", 18),
        ("Depth 3", 15), ("Depth 4", 12), ("Title", 25), ("Pre-condition", 20),
        ("Test Step", 45), ("Expected Result", 45), ("요구사항 ID", 12),
        ("Reference", 10), ("중요도", 8), ("Writer", 10),
    ]
    ROUND_SUBCOLUMNS = ["Result", "Severity", "Comments", "Issue #", "Tester"]
    NUM_TEST_ROUNDS = 3


# 기본 컬럼 매핑 (0-indexed)
DEFAULT_COLUMN_MAP = {
    "test_case_id": 1,      # B열
    "depth1": 2,            # C열
    "depth2": 3,            # D열
    "depth3": 4,            # E열
    "depth4": 5,            # F열
    "title": 6,             # G열
    "pre_condition": 7,     # H열
    "test_step": 8,         # I열
    "expected_result": 9,   # J열
    "requirement_id": 10,   # K열
    "reference": 11,        # L열
    "importance": 12,       # M열
    "writer": 13            # N열
}

# 테스트 결과 관련 컬럼 (1차, 2차, 3차...)
RESULT_COLUMNS = ["result", "defect_severity", "comments", "issue_number", "tester"]


def find_header_row(sheet, max_rows=20):
    """헤더 행 찾기 (Test Case ID가 있는 행)"""
    for row_idx in range(1, max_rows + 1):
        for col_idx in range(1, 20):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value and "test case id" in str(cell_value).lower():
                return row_idx
    return 10  # 기본값


def find_column_mapping(sheet, header_row):
    """헤더 행에서 컬럼 매핑 찾기"""
    column_map = {}

    for col_idx in range(1, 30):
        cell_value = sheet.cell(row=header_row, column=col_idx).value
        if not cell_value:
            continue

        cell_lower = str(cell_value).lower().strip()

        if "test case id" in cell_lower:
            column_map["test_case_id"] = col_idx
        elif cell_lower == "depth 1" or cell_lower == "depth1":
            column_map["depth1"] = col_idx
        elif cell_lower == "depth 2" or cell_lower == "depth2":
            column_map["depth2"] = col_idx
        elif cell_lower == "depth 3" or cell_lower == "depth3":
            column_map["depth3"] = col_idx
        elif cell_lower == "depth 4" or cell_lower == "depth4":
            column_map["depth4"] = col_idx
        elif "title" in cell_lower:
            column_map["title"] = col_idx
        elif "pre-condition" in cell_lower or "precondition" in cell_lower:
            column_map["pre_condition"] = col_idx
        elif "test step" in cell_lower:
            column_map["test_step"] = col_idx
        elif "expected" in cell_lower:
            column_map["expected_result"] = col_idx
        elif "요구사항" in str(cell_value) or "requirement" in cell_lower:
            column_map["requirement_id"] = col_idx
        elif "reference" in cell_lower:
            column_map["reference"] = col_idx
        elif "중요도" in str(cell_value) or "importance" in cell_lower:
            column_map["importance"] = col_idx
        elif "writer" in cell_lower:
            column_map["writer"] = col_idx

    return column_map


def copy_cell_style(source_cell, target_cell):
    """셀 스타일 복사"""
    if source_cell.font:
        target_cell.font = Font(
            name=source_cell.font.name,
            size=source_cell.font.size,
            bold=source_cell.font.bold,
            italic=source_cell.font.italic,
            color=source_cell.font.color
        )
    if source_cell.alignment:
        target_cell.alignment = Alignment(
            horizontal=source_cell.alignment.horizontal,
            vertical=source_cell.alignment.vertical,
            wrap_text=source_cell.alignment.wrap_text
        )
    if source_cell.border:
        target_cell.border = Border(
            left=source_cell.border.left,
            right=source_cell.border.right,
            top=source_cell.border.top,
            bottom=source_cell.border.bottom
        )
    if source_cell.fill:
        target_cell.fill = PatternFill(
            fill_type=source_cell.fill.fill_type,
            start_color=source_cell.fill.start_color,
            end_color=source_cell.fill.end_color
        )


def write_testcases_to_template(
    testcases_data: dict,
    template_path: Path,
    output_path: Path,
    sheet_name: str = None
):
    """테스트케이스를 템플릿에 작성"""

    # 템플릿 파일 복사
    shutil.copy(template_path, output_path)

    # 워크북 열기
    wb = load_workbook(output_path)

    # 시트 선택
    if sheet_name and sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active

    # 헤더 행 찾기
    header_row = find_header_row(sheet)
    print(f"Header row found at: {header_row}")

    # 컬럼 매핑 찾기
    column_map = find_column_mapping(sheet, header_row)
    print(f"Column mapping: {column_map}")

    # 데이터 시작 행 (헤더 다음 행)
    data_start_row = header_row + 1

    # 기존 데이터 행에서 스타일 가져오기 (있는 경우)
    style_source_row = data_start_row
    if sheet.cell(row=style_source_row, column=1).value is None:
        style_source_row = header_row  # 데이터가 없으면 헤더 스타일 사용

    testcases = testcases_data.get("testcases", [])

    for idx, tc in enumerate(testcases):
        row_idx = data_start_row + idx

        for field, col_idx in column_map.items():
            value = tc.get(field, "")
            cell = sheet.cell(row=row_idx, column=col_idx)
            cell.value = value

            # 스타일 복사 (첫 번째 데이터 행의 스타일 기준)
            style_cell = sheet.cell(row=style_source_row, column=col_idx)
            copy_cell_style(style_cell, cell)

            # 줄바꿈이 있는 필드는 wrap_text 활성화
            if value and "\n" in str(value):
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal if cell.alignment else "left",
                    vertical="top",
                    wrap_text=True
                )

    # 저장
    wb.save(output_path)
    print(f"Saved {len(testcases)} test cases to: {output_path}")

    return output_path


def create_new_testcase_excel(testcases_data: dict, output_path: Path):
    """템플릿 없이 새 Excel 파일 생성 (테스트 회차 포함)"""
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Test Cases"

    # 프로젝트 정보
    project_info = testcases_data.get("project_info", {})
    project_name = project_info.get("project_name", "Test Project")
    version = project_info.get("version", "v1.0")
    total_tc = testcases_data.get("total_testcases", 0)

    # 색상 정의 (중앙 설정에서 가져옴)
    header_fill = PatternFill(start_color=EXCEL_COLORS["header"], end_color=EXCEL_COLORS["header"], fill_type="solid")
    subheader_fill = PatternFill(start_color=EXCEL_COLORS["subheader"], end_color=EXCEL_COLORS["subheader"], fill_type="solid")
    round1_fill = PatternFill(start_color=EXCEL_COLORS["round1"], end_color=EXCEL_COLORS["round1"], fill_type="solid")
    round2_fill = PatternFill(start_color=EXCEL_COLORS["round2"], end_color=EXCEL_COLORS["round2"], fill_type="solid")
    round3_fill = PatternFill(start_color=EXCEL_COLORS["round3"], end_color=EXCEL_COLORS["round3"], fill_type="solid")
    summary_fill = PatternFill(start_color=EXCEL_COLORS["summary"], end_color=EXCEL_COLORS["summary"], fill_type="solid")
    blank_fill = PatternFill(start_color=EXCEL_COLORS["blank_field"], end_color=EXCEL_COLORS["blank_field"], fill_type="solid")

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ===== 상단 요약 영역 (행 1~5) =====
    # 제목
    sheet.merge_cells("A1:N1")
    title_cell = sheet["A1"]
    title_cell.value = f"{project_name} - 테스트케이스"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 30

    # 요약 정보
    summary_labels = [
        ("A3", "Version"), ("B3", version),
        ("C3", "Total TC"), ("D3", total_tc),
        ("E3", "Pass"), ("F3", "=COUNTIF(O:O,\"Pass\")+COUNTIF(T:T,\"Pass\")+COUNTIF(Y:Y,\"Pass\")"),
        ("G3", "Fail"), ("H3", "=COUNTIF(O:O,\"Fail\")+COUNTIF(T:T,\"Fail\")+COUNTIF(Y:Y,\"Fail\")"),
        ("I3", "N/T"), ("J3", f"={total_tc}-F3-H3"),
    ]
    for cell_ref, value in summary_labels:
        cell = sheet[cell_ref]
        cell.value = value
        cell.border = thin_border
        cell.fill = summary_fill
        if cell_ref[0] in "ACEGI":
            cell.font = Font(bold=True)

    # ===== 헤더 행 1 (병합 헤더) - 행 5 =====
    header_row1 = 5
    header_row2 = 6

    # 기본 컬럼 (중앙 설정에서 가져옴)
    base_headers = []
    for idx, (name, width) in enumerate(BASE_COLUMNS):
        col_letter = get_column_letter(idx + 1)
        base_headers.append((col_letter, name, width))

    # 테스트 회차 컬럼 (중앙 설정에서 가져옴)
    round_headers = ROUND_SUBCOLUMNS

    # 기본 헤더 작성 (병합)
    for col_letter, header_text, width in base_headers:
        col_idx = ord(col_letter) - ord('A') + 1
        # 행 5, 6 병합
        sheet.merge_cells(start_row=header_row1, start_column=col_idx,
                          end_row=header_row2, end_column=col_idx)
        cell = sheet.cell(row=header_row1, column=col_idx)
        cell.value = header_text
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        # 병합된 하단 셀에도 테두리
        sheet.cell(row=header_row2, column=col_idx).border = thin_border
        sheet.column_dimensions[col_letter].width = width

    # 테스트 회차 헤더 작성
    round_fills = [round1_fill, round2_fill, round3_fill]
    round_start_col = len(base_headers) + 1  # O열부터

    for round_num in range(1, NUM_TEST_ROUNDS + 1):  # 1차, 2차, 3차
        round_fill = round_fills[round_num - 1]
        start_col = round_start_col + (round_num - 1) * 5

        # 상위 헤더 (회차명) - 5개 컬럼 병합
        sheet.merge_cells(start_row=header_row1, start_column=start_col,
                          end_row=header_row1, end_column=start_col + 4)
        round_cell = sheet.cell(row=header_row1, column=start_col)
        round_cell.value = f"{round_num}차 테스트"
        round_cell.font = Font(bold=True, color="FFFFFF")
        round_cell.fill = round_fill
        round_cell.alignment = header_alignment
        round_cell.border = thin_border

        # 하위 헤더 (세부 항목)
        for sub_idx, sub_header in enumerate(round_headers):
            col = start_col + sub_idx
            cell = sheet.cell(row=header_row2, column=col)
            cell.value = sub_header
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = round_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            # 컬럼 너비
            col_letter = get_column_letter(col)
            if sub_header == "Result":
                sheet.column_dimensions[col_letter].width = 8
            elif sub_header == "Severity":
                sheet.column_dimensions[col_letter].width = 10
            elif sub_header == "Comments":
                sheet.column_dimensions[col_letter].width = 20
            elif sub_header == "Issue #":
                sheet.column_dimensions[col_letter].width = 10
            else:
                sheet.column_dimensions[col_letter].width = 10

    # 행 높이 설정
    sheet.row_dimensions[header_row1].height = 25
    sheet.row_dimensions[header_row2].height = 20

    # ===== 데이터 작성 =====
    data_alignment = Alignment(vertical="top", wrap_text=True)
    testcases = testcases_data.get("testcases", [])
    data_start_row = header_row2 + 1  # 7행부터

    # 총 컬럼 수 (기본 컬럼 + 회차별 컬럼)
    total_cols = len(base_headers) + len(ROUND_SUBCOLUMNS) * NUM_TEST_ROUNDS

    def safe_cell_value(value):
        """openpyxl에서 불법 문자 제거 및 안전한 셀 값 생성

        openpyxl은 XML에서 허용되지 않는 제어 문자를 거부함
        - ASCII 0x00-0x08, 0x0B-0x0C, 0x0E-0x1F (탭, 줄바꿈 제외)
        이러한 문자를 제거하여 안전하게 처리
        """
        import re
        if not value:
            return value
        str_val = str(value)
        # XML에서 허용되지 않는 제어 문자 제거 (탭 \x09, 줄바꿈 \x0A, 캐리지리턴 \x0D 제외)
        # 허용되지 않는 범위: \x00-\x08, \x0B-\x0C, \x0E-\x1F
        illegal_xml_chars = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
        clean_val = illegal_xml_chars.sub('', str_val)
        return clean_val

    # 공란 처리 대상 필드 (필드명, 컬럼 인덱스)
    blank_check_fields = [
        ("pre_condition", 8),
        ("test_step", 9),
        ("expected_result", 10),
    ]

    for idx, tc in enumerate(testcases, start=1):
        row = data_start_row + idx - 1

        # 기본 데이터 (safe_cell_value로 예약어 처리)
        sheet.cell(row=row, column=1).value = idx
        sheet.cell(row=row, column=2).value = tc.get("test_case_id", "")
        sheet.cell(row=row, column=3).value = tc.get("depth1", "")
        sheet.cell(row=row, column=4).value = tc.get("depth2", "")
        sheet.cell(row=row, column=5).value = tc.get("depth3", "")
        sheet.cell(row=row, column=6).value = tc.get("depth4", "")
        sheet.cell(row=row, column=7).value = tc.get("title", "")
        sheet.cell(row=row, column=8).value = tc.get("pre_condition", "")
        sheet.cell(row=row, column=9).value = safe_cell_value(tc.get("test_step", ""))
        sheet.cell(row=row, column=10).value = safe_cell_value(tc.get("expected_result", ""))
        sheet.cell(row=row, column=11).value = tc.get("requirement_id", "")
        sheet.cell(row=row, column=12).value = tc.get("reference", "")
        sheet.cell(row=row, column=13).value = tc.get("importance", "")
        sheet.cell(row=row, column=14).value = tc.get("writer", "")

        # 공란 필드에 노란색 배경 + 코멘트 적용
        blank_reasons = tc.get("_blank_reasons", {})
        for field_name, col_idx in blank_check_fields:
            value = tc.get(field_name, "")
            # 값이 비어있고 _blank_reasons에 해당 필드가 있으면 노란색 + 코멘트 적용
            if not value and field_name in blank_reasons:
                cell = sheet.cell(row=row, column=col_idx)
                cell.fill = blank_fill
                comment = Comment(
                    text=blank_reasons[field_name],
                    author="TC Generator"
                )
                cell.comment = comment

        # 모든 셀에 스타일 적용
        for col in range(1, total_cols + 1):
            cell = sheet.cell(row=row, column=col)
            cell.alignment = data_alignment
            cell.border = thin_border

    # 필터 설정 (헤더 행2부터)
    last_row = data_start_row + len(testcases) - 1 if testcases else data_start_row
    last_col_letter = get_column_letter(total_cols)
    sheet.auto_filter.ref = f"A{header_row2}:{last_col_letter}{last_row}"

    # 헤더 행 고정
    sheet.freeze_panes = f"A{data_start_row}"

    # 데이터 검증 (Result 컬럼에 드롭다운 - 중앙 설정에서 옵션 가져옴)
    from openpyxl.worksheet.datavalidation import DataValidation
    result_options_str = ",".join(RESULT_OPTIONS)
    result_validation = DataValidation(
        type="list",
        formula1=f'"{result_options_str}"',
        allow_blank=True
    )
    result_validation.error = "유효한 결과값을 선택하세요"
    result_validation.errorTitle = "Invalid Result"

    severity_options_str = ",".join(SEVERITY_OPTIONS)
    severity_validation = DataValidation(
        type="list",
        formula1=f'"{severity_options_str}"',
        allow_blank=True
    )

    sheet.add_data_validation(result_validation)
    sheet.add_data_validation(severity_validation)

    # Result 컬럼에 validation 적용 (O, T, Y열 등)
    for round_num in range(NUM_TEST_ROUNDS):
        result_col = round_start_col + round_num * 5
        severity_col = result_col + 1
        result_col_letter = get_column_letter(result_col)
        severity_col_letter = get_column_letter(severity_col)

        if testcases:
            result_validation.add(f"{result_col_letter}{data_start_row}:{result_col_letter}{last_row}")
            severity_validation.add(f"{severity_col_letter}{data_start_row}:{severity_col_letter}{last_row}")

    wb.save(output_path)
    print(f"Created new Excel file with {len(testcases)} test cases: {output_path}")
    print(f"  - 1차/2차/3차 테스트 회차 컬럼 포함")
    print(f"  - 상단 요약 영역 포함")

    return output_path


def main():
    if len(sys.argv) < 3:
        print("Usage: python write_excel.py <testcases_json> <output_xlsx> [template_xlsx]")
        sys.exit(1)

    testcases_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    template_path = Path(sys.argv[3]) if len(sys.argv) > 3 else None

    if not testcases_path.exists():
        print(f"Error: File not found: {testcases_path}")
        sys.exit(1)

    with open(testcases_path, "r", encoding="utf-8") as f:
        testcases_data = json.load(f)

    if template_path and template_path.exists():
        write_testcases_to_template(testcases_data, template_path, output_path)
    else:
        create_new_testcase_excel(testcases_data, output_path)

    return output_path


if __name__ == "__main__":
    main()
